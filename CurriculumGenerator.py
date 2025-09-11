#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
合并版课程表生成器
整合了原版、修复版和优化版的所有功能和改进
- 修复了学期开始日期处理问题
- 优化了课程信息解析逻辑
- 增强了错误处理和日志输出
- 支持多种课程格式和特殊情况处理
"""

import openpyxl
import sys
import Curriculum
import re
import datetime
import requests
import json

# 课程时间映射
__course_start_time = {
    "1": datetime.time(8),
    "3": datetime.time(10, 20),
    "5": datetime.time(14),
    "7": datetime.time(16, 20),
    "9": datetime.time(19)
}

__course_end_time = {
    "2": datetime.time(9, 50),
    "4": datetime.time(12, 10),
    "6": datetime.time(15, 50),
    "8": datetime.time(18, 10),
    "10": datetime.time(20, 50),
    "11": datetime.time(21, 50)
}

def usage():
    """显示使用说明"""
    print("用法: python3 CurriculumGenerator_merged.py <Excel文件> <学期开始日期> <学期结束日期> [路程时间]")
    print("示例: python3 CurriculumGenerator_merged.py export.xlsx 20250908 20251228 30")
    print("日期格式: YYYYMMDD")
    print("路程时间: 可选参数，单位为分钟，默认30分钟")
    print("注意: 学期开始时间为学期第一周的周一，学期结束时间为学期最后一周的周末")
    sys.exit(1)

def get_holidays_and_workdays(start_date, end_date):
    """获取节假日和调休工作日信息"""
    print("正在获取节假日和调休工作日信息...")
    holidays = []
    workdays = []  # 调休工作日
    
    try:
        # 尝试使用范围API获取节假日信息
        start_str = start_date.strftime("%Y%m%d")
        end_str = end_date.strftime("%Y%m%d")
        
        url = f"https://timor.tech/api/holiday/range/{start_str}/{end_str}"
        response = requests.get(url, timeout=10)
        data = response.json()
        
        if data["code"] == 0:
            holiday_data = data["holiday"]
            for date_str, info in holiday_data.items():
                year, month, day = map(int, date_str.split("-"))
                date_obj = datetime.date(year, month, day)
                if info["holiday"]:
                    holidays.append(date_obj)
                elif info.get("work", False):  # 调休工作日
                    workdays.append(date_obj)
            print(f"成功获取到{len(holidays)}个节假日，{len(workdays)}个调休工作日")
            return holidays, workdays
        else:
            print(f"API返回错误: {data['msg']}")
    except Exception as e:
        print(f"获取节假日信息出错: {str(e)}")
    
    # 如果API失败，尝试年度API
    try:
        year = start_date.year
        url = f"http://timor.tech/api/holiday/year/{year}"
        response = requests.get(url, timeout=5)
        
        if response.status_code == 200:
            data = response.json()
            for date_str, info in data.get('holiday', {}).items():
                try:
                    date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
                    if start_date <= date_obj <= end_date:
                        if info.get('holiday', False):
                            holidays.append(date_obj)
                        elif info.get('work', False):  # 调休工作日
                            workdays.append(date_obj)
                except ValueError:
                    continue
            print(f"从年度API获取到{len(holidays)}个节假日，{len(workdays)}个调休工作日")
            return holidays, workdays
    except Exception as e:
        print(f"年度API获取失败: {str(e)}")
    
    # 使用默认节假日信息
    print("将使用默认节假日和调休工作日信息")
    current_year = start_date.year
    holidays, workdays = get_default_holidays_and_workdays(current_year)
    print(f"使用默认信息，共{len(holidays)}个节假日，{len(workdays)}个调休工作日")
    return holidays, workdays

def get_default_holidays_and_workdays(year):
    """获取默认节假日和调休工作日信息"""
    if year == 2025:
        holidays = [
            # 元旦
            datetime.date(2025, 1, 1),
            # 春节
            datetime.date(2025, 1, 29), datetime.date(2025, 1, 30), datetime.date(2025, 1, 31),
            datetime.date(2025, 2, 1), datetime.date(2025, 2, 2), datetime.date(2025, 2, 3), datetime.date(2025, 2, 4),
            # 清明节
            datetime.date(2025, 4, 5), datetime.date(2025, 4, 6), datetime.date(2025, 4, 7),
            # 劳动节
            datetime.date(2025, 5, 1), datetime.date(2025, 5, 2), datetime.date(2025, 5, 3), datetime.date(2025, 5, 4), datetime.date(2025, 5, 5),
            # 端午节
            datetime.date(2025, 6, 2), datetime.date(2025, 6, 3), datetime.date(2025, 6, 4),
            # 中秋节
            datetime.date(2025, 9, 13), datetime.date(2025, 9, 14), datetime.date(2025, 9, 15),
            # 国庆节
            datetime.date(2025, 10, 1), datetime.date(2025, 10, 2), datetime.date(2025, 10, 3),
            datetime.date(2025, 10, 4), datetime.date(2025, 10, 5), datetime.date(2025, 10, 6), datetime.date(2025, 10, 7)
        ]
        workdays = [
            # 春节调休工作日
            datetime.date(2025, 1, 26),  # 周日调休
            datetime.date(2025, 2, 8),   # 周六调休
            # 劳动节调休工作日
            datetime.date(2025, 4, 27),  # 周日调休
            # 国庆节调休工作日
            datetime.date(2025, 9, 28),  # 周日调休
            datetime.date(2025, 10, 11)  # 周六调休
        ]
        return holidays, workdays
    elif year == 2024:
        holidays = [
            # 元旦
            datetime.date(2024, 1, 1),
            # 春节
            datetime.date(2024, 2, 10), datetime.date(2024, 2, 11), datetime.date(2024, 2, 12),
            datetime.date(2024, 2, 13), datetime.date(2024, 2, 14), datetime.date(2024, 2, 15), datetime.date(2024, 2, 16), datetime.date(2024, 2, 17),
            # 清明节
            datetime.date(2024, 4, 4), datetime.date(2024, 4, 5), datetime.date(2024, 4, 6),
            # 劳动节
            datetime.date(2024, 5, 1), datetime.date(2024, 5, 2), datetime.date(2024, 5, 3), datetime.date(2024, 5, 4), datetime.date(2024, 5, 5),
            # 端午节
            datetime.date(2024, 6, 10),
            # 中秋节
            datetime.date(2024, 9, 15), datetime.date(2024, 9, 16), datetime.date(2024, 9, 17),
            # 国庆节
            datetime.date(2024, 10, 1), datetime.date(2024, 10, 2), datetime.date(2024, 10, 3),
            datetime.date(2024, 10, 4), datetime.date(2024, 10, 5), datetime.date(2024, 10, 6), datetime.date(2024, 10, 7)
        ]
        workdays = [
            # 春节调休工作日
            datetime.date(2024, 2, 4),   # 周日调休
            datetime.date(2024, 2, 18),  # 周日调休
            # 劳动节调休工作日
            datetime.date(2024, 4, 28),  # 周日调休
            datetime.date(2024, 5, 11),  # 周六调休
            # 国庆节调休工作日
            datetime.date(2024, 9, 29),  # 周日调休
            datetime.date(2024, 10, 12)  # 周六调休
        ]
        return holidays, workdays
    else:
        print(f"警告：没有{year}年的预设节假日信息，将使用2024年的节假日信息")
        return get_default_holidays_and_workdays(2024)

def extract_course_info(course_text):
    """从课程文本中提取结构化信息"""
    courses = []
    lines = course_text.split('\n')
    
    # 每4行为一个课程信息块
    for i in range(0, len(lines), 4):
        if i + 3 < len(lines):
            course_name = lines[i].strip()
            teacher = lines[i + 1].strip() if i + 1 < len(lines) else ""
            class_info = lines[i + 2].strip() if i + 2 < len(lines) else ""
            time_location = lines[i + 3].strip() if i + 3 < len(lines) else ""
            
            # 跳过实验室安全学，由process_special_course处理
            if "实验室安全学" in course_name:
                continue
                
            if course_name and '[' in time_location:
                # 解析周次信息
                week_pattern = re.findall(r'\[(\d+)-(\d+)(单|双)?周\]', time_location)
                if not week_pattern:
                    week_pattern = re.findall(r'\[(\d+)周\]', time_location)
                    if week_pattern:
                        week_pattern = [(week_pattern[0], week_pattern[0], "")]
                
                # 解析地点信息
                location_pattern = re.findall(r'\]\[([^\]]+)\]\[', time_location)
                location = location_pattern[0] if location_pattern else "未知地点"
                
                # 解析时间段信息
                time_pattern = re.findall(r'\[(\d+)-(\d+)节\]', time_location)
                
                courses.append({
                    'name': course_name,
                    'teacher': teacher,
                    'class_info': class_info,
                    'week_info': week_pattern[0] if week_pattern else None,
                    'location': location,
                    'time_slots': time_pattern[0] if time_pattern else None,
                    'raw_text': time_location
                })
    
    return courses

def add_course_to_curriculum(curriculum, course_info, day_offset, term_start_date, term_end_date, holidays, default_travel_time, is_single_event=False, workdays=None):
    """将课程信息添加到课程表中"""
    if not course_info['week_info'] or not course_info['time_slots']:
        print(f"警告: 课程 {course_info['name']} 缺少必要信息，跳过")
        return
    
    start_week, end_week, week_type = course_info['week_info']
    start_time_slot, end_time_slot = course_info['time_slots']
    
    # 计算具体的开始和结束时间
    course_start_time = __course_start_time.get(start_time_slot, datetime.time(8))
    course_end_time = __course_end_time.get(end_time_slot, datetime.time(9, 50))
    
    # 设置地点
    if course_info['location'] == "无地点" or course_info['location'] == "未知地点":
        location = "南方科技大学-在线课程"
    else:
        location = f"南方科技大学-{course_info['location']}"
    
    # 设置提醒时间
    travel_time = default_travel_time
    if "实验室安全学" in course_info['name']:
        travel_time = default_travel_time * 2
    elif "实验" in course_info['name'] or "实践" in course_info['name']:
        travel_time = int(default_travel_time * 1.5)
    elif "第一科研楼" in location or "荔园" in location:
        travel_time = int(default_travel_time * 1.33)
    
    # 循环遍历周次范围
    for w in range(int(start_week), int(end_week) + 1):
        if week_type == "单" and w % 2 == 0:
            continue
        if week_type == "双" and w % 2 == 1:
            continue
        if week_type not in ["单", "双", ""] :
            continue  # 如果有其他类型，跳过（可选扩展）
        
        course_date = term_start_date + datetime.timedelta(days=(w - 1) * 7 + day_offset)
        if course_date > term_end_date:
            continue
        
        if course_date in holidays:
            print(f"跳过节假日课程: {course_info['name']} ({course_date})")
            continue
        
        weekday = course_date.weekday()
        if weekday >= 5 and (workdays is None or course_date not in workdays):
            print(f"跳过周末课程: {course_info['name']} ({course_date})")
            continue
        
        start_datetime = datetime.datetime.combine(course_date, course_start_time)
        end_datetime = datetime.datetime.combine(course_date, course_end_time)
        
        # 添加单个事件
        single_event_end = datetime.datetime.combine(course_date, datetime.time.max)
        Curriculum.add_course(curriculum, course_info['name'], start_datetime, end_datetime, 
                             location, Curriculum.CourseRepetitionType.weekly, single_event_end, 
                             travel_time_minutes=travel_time, is_single_event=True)
        
        print(f"添加课程: {course_info['name']} - {start_datetime.strftime('%Y-%m-%d %H:%M')} ({location})")

def process_special_course(course_value, sheet, course_row, day, term_start_date, term_end_date, holidays, default_travel_time, curriculum, workdays=None):
    """处理特殊格式的课程（如实验室安全学）"""
    if "实验室安全学" in course_value:
        name = "实验室安全学"
        
        # 确定地点
        if "[无地点]" in course_value:
            location = "南方科技大学-在线课程"
        elif "一科报告厅" in course_value:
            location = "南方科技大学-第一科研楼报告厅"
        else:
            location = "南方科技大学-第一科研楼101"
        
        # 获取时间段信息
        row_header = sheet[f"A{course_row}"].value
        if row_header and "第" in str(row_header) and "节" in str(row_header):
            num_str = re.findall(r"\d+", str(row_header))
            if len(num_str) >= 2:
                num = [num_str[0], num_str[1]]
            else:
                num = [num_str[0], num_str[0]] if len(num_str) > 0 else ["1", "2"]
        else:
            num = ["1", "2"]  # 默认值
        
        # 解析周次信息
        week_patterns = re.findall(r"\[(\d+)-(\d+)(单|双)?周\]", course_value)
        single_week_patterns = re.findall(r"\[(\d+)周\]", course_value)
        
        if week_patterns:
            # 处理范围周次，如[2-12双周]
            start_week, end_week, week_type = week_patterns[0]
            
            # 循环遍历周次范围
            for w in range(int(start_week), int(end_week) + 1):
                if week_type == "单" and w % 2 == 0:
                    continue
                if week_type == "双" and w % 2 == 1:
                    continue
                
                course_date = term_start_date + datetime.timedelta(days=(w - 1) * 7 + day)
                if course_date > term_end_date:
                    continue
                
                if course_date in holidays:
                    print(f"跳过节假日课程: {name} ({course_date})")
                    continue
                
                weekday = course_date.weekday()
                if weekday >= 5 and (workdays is None or course_date not in workdays):
                    print(f"跳过周末课程: {name} ({course_date})")
                    continue
                
                start_time = datetime.datetime.combine(course_date, __course_start_time.get(num[0]))
                end_time = datetime.datetime.combine(course_date, __course_end_time.get(num[1]))
                
                # 添加单个事件
                single_event_end = datetime.datetime.combine(course_date, datetime.time.max)
                safety_travel_time = default_travel_time * 2
                Curriculum.add_course(curriculum, name, start_time, end_time, location, Curriculum.CourseRepetitionType.weekly, single_event_end, travel_time_minutes=safety_travel_time, is_single_event=True)
                print(f"添加实验室安全学: {start_time.strftime('%Y-%m-%d %H:%M')} - {end_time.strftime('%H:%M')} @ {location}")
        
        elif single_week_patterns:
            # 处理单次课程，如[12周]
            target_week = int(single_week_patterns[0])
            
            # 解析时间段信息 [1-8节]
            time_slot_patterns = re.findall(r"\[(\d+)-(\d+)节\]", course_value)
            if time_slot_patterns:
                start_slot, end_slot = time_slot_patterns[0]
                
                # 计算目标周的日期
                target_date = term_start_date + datetime.timedelta(days=(target_week-1)*7 + day)
                
                # 为每个时间段创建单独的课程事件
                current_slot = int(start_slot)
                while current_slot <= int(end_slot):
                    # 确定当前时间段的开始和结束时间
                    if current_slot % 2 == 1:  # 奇数节为开始时间
                        slot_start_time = __course_start_time.get(str(current_slot), datetime.time(8))
                        if current_slot < int(end_slot):
                            slot_end_time = __course_end_time.get(str(current_slot + 1), datetime.time(9, 50))
                            current_slot += 2
                        else:
                            slot_end_time = __course_end_time.get(str(current_slot + 1), datetime.time(9, 50))
                            current_slot += 1
                    else:
                        current_slot += 1
                        continue
                    
                    start_datetime = datetime.datetime.combine(target_date, slot_start_time)
                    end_datetime = datetime.datetime.combine(target_date, slot_end_time)
                    
                    # 检查是否为节假日或调休工作日
                    course_date = start_datetime.date()
                    weekday = course_date.weekday()
                    
                    # 实验室安全学单次课程特殊处理：只跳过节假日，不跳过周末
                    # 其他课程：节假日跳过，周末非调休日跳过，调休工作日正常上课
                    if (course_date not in holidays):
                        safety_travel_time = default_travel_time * 2
                        # 单次课程不需要重复，设置结束时间为当天
                        single_event_end = datetime.datetime.combine(target_date, datetime.time.max)
                        Curriculum.add_course(curriculum, name, start_datetime, end_datetime, location, 
                                             Curriculum.CourseRepetitionType.weekly, single_event_end, 
                                             travel_time_minutes=safety_travel_time, is_single_event=True)
                        print(f"添加实验室安全学(单次): {start_datetime.strftime('%Y-%m-%d %H:%M')} - {end_datetime.strftime('%H:%M')} @ {location}")
            else:
                # 如果没有找到时间段信息，使用默认时间
                target_date = term_start_date + datetime.timedelta(days=(target_week-1)*7 + day)
                start_time = datetime.datetime.combine(target_date, __course_start_time.get(num[0]))
                end_time = datetime.datetime.combine(target_date, __course_end_time.get(num[1]))
                
                course_date = start_time.date()
                weekday = course_date.weekday()
                
                # 实验室安全学单次课程特殊处理：只跳过节假日，不跳过周末
                if (course_date not in holidays):
                    safety_travel_time = default_travel_time * 2
                    single_event_end = datetime.datetime.combine(target_date, datetime.time.max)
                    Curriculum.add_course(curriculum, name, start_time, end_time, location, 
                                         Curriculum.CourseRepetitionType.weekly, single_event_end, 
                                         travel_time_minutes=safety_travel_time, is_single_event=True)
                    print(f"添加实验室安全学(单次): {start_time.strftime('%Y-%m-%d %H:%M')} - {end_time.strftime('%H:%M')} @ {location}")
        return True
    return False

def process_regular_course(course_value, sheet, course_row, day, term_start_date, term_end_date, holidays, default_travel_time, curriculum, workdays=None):
    """处理常规格式的课程"""
    val = course_value.split("\n")
    
    # 尝试使用优化版的结构化解析
    structured_courses = extract_course_info(course_value)
    if structured_courses:
        for course_info in structured_courses:
            add_course_to_curriculum(curriculum, course_info, day, term_start_date, term_end_date, holidays, default_travel_time, workdays=workdays)
        return
    
    # 如果结构化解析失败，使用原版的解析方法
    for i in range(len(val)//4):
        name = val[i * 4]
        if not name.strip():
            continue
            
        info = val[i * 4 + 3].split("][")
        
        # 获取时间段信息
        row_header = sheet[f"A{course_row}"].value
        if row_header and "第" in str(row_header) and "节" in str(row_header):
            num = re.findall(r"\d+", str(row_header))
        else:
            # 从课程信息中提取时间段
            time_info = re.findall(r"\[(\d+)-(\d+)节\]", course_value)
            if time_info:
                num = [time_info[0][0], time_info[0][1]]
            else:
                num = ["1", "2"]  # 默认值
        
        # 解析周次类型和时间
        if len(info) > 0 and len(info[0]) > 2:
            if info[0][-2] == '单':
                repeat_type = Curriculum.CourseRepetitionType.biweekly
                start_time = datetime.datetime.combine(term_start_date + datetime.timedelta(days=day), __course_start_time.get(num[0]))
                end_time = datetime.datetime.combine(term_start_date + datetime.timedelta(days=day), __course_end_time.get(num[1]))
            elif info[0][-2] == '双':
                repeat_type = Curriculum.CourseRepetitionType.biweekly
                start_time = datetime.datetime.combine(term_start_date + datetime.timedelta(days=day+7), __course_start_time.get(num[0]))
                end_time = datetime.datetime.combine(term_start_date + datetime.timedelta(days=day+7), __course_end_time.get(num[1]))
            else:
                repeat_type = Curriculum.CourseRepetitionType.weekly
                start_time = datetime.datetime.combine(term_start_date + datetime.timedelta(days=day), __course_start_time.get(num[0]))
                end_time = datetime.datetime.combine(term_start_date + datetime.timedelta(days=day), __course_end_time.get(num[1]))
        else:
            repeat_type = Curriculum.CourseRepetitionType.weekly
            start_time = datetime.datetime.combine(term_start_date + datetime.timedelta(days=day), __course_start_time.get(num[0]))
            end_time = datetime.datetime.combine(term_start_date + datetime.timedelta(days=day), __course_end_time.get(num[1]))
        
        # 确定地点
        if len(info) > 1:
            location = "南方科技大学-" + info[1]
        else:
            location = "南方科技大学-未知地点"
        
        # 检查是否为节假日或调休工作日
        course_date = start_time.date()
        weekday = course_date.weekday()
        
        # 节假日跳过，周末非调休日跳过，调休工作日正常上课
        if (course_date in holidays or 
            (weekday >= 5 and (not workdays or course_date not in workdays))):
            print(f"跳过节假日或非工作日课程: {name} on {course_date}")
            continue
        
        # 根据课程类型调整路程时间
        travel_time = default_travel_time
        if "实验" in name or "实践" in name:
            travel_time = int(default_travel_time * 1.5)
        elif "第一科研楼" in location or "荔园" in location:
            travel_time = int(default_travel_time * 1.33)
        
        term_end_datetime = datetime.datetime.combine(term_end_date, datetime.time.max)
        Curriculum.add_course(curriculum, name, start_time, end_time, location, repeat_type, term_end_datetime, travel_time_minutes=travel_time)
        print(f"添加课程: {name} - {start_time.strftime('%Y-%m-%d %H:%M')} - {end_time.strftime('%H:%M')} @ {location}")

def main():
    """主函数"""
    # 参数检查
    if len(sys.argv) < 4 or len(sys.argv) > 5:
        usage()
    
    excel_file = sys.argv[1]
    if not excel_file.endswith(".xlsx"):
        print("错误：请提供.xlsx格式的Excel文件")
        usage()
    
    try:
        term_start_str = sys.argv[2]
        term_end_str = sys.argv[3]
        
        if len(term_start_str) != 8 or len(term_end_str) != 8:
            raise ValueError("日期格式错误")
        
        term_start_date = datetime.date(int(term_start_str[0:4]), int(term_start_str[4:6]), int(term_start_str[6:]))
        term_end_date = datetime.date(int(term_end_str[0:4]), int(term_end_str[4:6]), int(term_end_str[6:]))
    except ValueError:
        print("错误：日期格式必须为YYYYMMDD")
        usage()
    
    # 路程时间设置
    default_travel_time = 30
    if len(sys.argv) == 5:
        try:
            default_travel_time = int(sys.argv[4])
            if default_travel_time < 0:
                print("错误：路程时间提醒必须是正整数")
                usage()
        except ValueError:
            print("错误：路程时间提醒必须是整数")
            usage()
    
    print(f"开始处理课程表: {excel_file}")
    print(f"学期时间: {term_start_date} 到 {term_end_date}")
    print(f"路程提醒时间: {default_travel_time} 分钟")
    print()
    
    # 加载Excel文件
    try:
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
    except Exception as e:
        print(f"错误：无法读取Excel文件 - {str(e)}")
        sys.exit(1)
    
    # 获取节假日和调休工作日信息
    holidays, workdays = get_holidays_and_workdays(term_start_date, term_end_date)
    
    # 创建课程表对象
    curriculum = Curriculum.Curriculum()
    
    print("开始解析课程信息...")
    
    # 遍历Excel表格的列（星期一到星期日）
    day = 0  # 0=星期一, 1=星期二, ..., 6=星期日
    total_courses = 0
    
    for col in sheet.iter_cols(min_col=2, max_col=8, min_row=4, max_row=sheet.max_row):
        print(f"\n处理第{day+1}列 (星期{['一', '二', '三', '四', '五', '六', '日'][day]}):")
        
        for course in col:
            if course.value and isinstance(course.value, str):
                print(f"  第{course.row}行: {repr(course.value[:50])}...")
                
                # 首先尝试处理特殊课程
                if process_special_course(course.value, sheet, course.row, day, term_start_date, term_end_date, holidays, default_travel_time, curriculum, workdays):
                    total_courses += 1
                    continue
                
                # 处理常规课程
                if '[' in course.value:
                    process_regular_course(course.value, sheet, course.row, day, term_start_date, term_end_date, holidays, default_travel_time, curriculum, workdays)
                    total_courses += 1
        
        day += 1
    
    print(f"\n总共处理了 {total_courses} 个课程时间段")
    
    # 保存ICS文件
    curriculum.save_as_ics_file()
    print(f"课程表已保存为: {curriculum.calendar_name}.ics")
    
    # 统计生成的课程
    ics_content = curriculum.get_ics_text()
    event_count = ics_content.count('BEGIN:VEVENT')
    print(f"生成的日历事件数量: {event_count}")

if __name__ == "__main__":
    main()