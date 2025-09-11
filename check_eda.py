import openpyxl

wb = openpyxl.load_workbook('export.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
    for cell in row:
        if cell.value and '先进电子设计自动化EDA' in str(cell.value):
            print('Row {}, Column {}: {}'.format(cell.row, cell.column_letter, cell.value))